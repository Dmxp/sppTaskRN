from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def build_excel_file(calculation: dict) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SPP calculation"

    headers = [
        "№",
        "Код СПП",
        "Наименование",
        "Уровень",
        "Сумма",
        "Отделы"
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    tree = calculation["result_json"]["tree"]

    row_number = 1

    def walk(nodes, level, prefix):
        nonlocal row_number

        for index, node in enumerate(nodes, start=1):
            number = f"{prefix}.{index}" if prefix else str(index)

            departments = ", ".join(
                department["name"]
                for department in node.get("departments", [])
            )

            sheet.append([
                number,
                node["code"],
                node["name"],
                level,
                node.get("amount", 0),
                departments
            ])

            row_number += 1

            children = node.get("children", [])

            if children:
                walk(children, level + 1, number)

    walk(tree, 1, "")

    column_widths = {
        "A": 12,
        "B": 16,
        "C": 36,
        "D": 12,
        "E": 16,
        "F": 28
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True
            )

    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)

    return file_stream